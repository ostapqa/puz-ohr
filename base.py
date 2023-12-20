import os
import pprint
import random
from statistics import mean
from math import sqrt

import openpyxl
import pandas as pd

number_of_experts = 10
filename = "table.xlsx"


def get_sheet(filepath, index):
    wb = openpyxl.load_workbook(filepath)

    sheet_names = wb.sheetnames

    sheet = wb[sheet_names[index - 1]]
    df = pd.DataFrame(sheet.values)

    return df


def select_random_experts():
    numbers = list(range(1, 1001))
    random_numbers = random.sample(numbers, number_of_experts)

    sorted_keys = [f"E{i}" for i in range(1, number_of_experts)]

    result = dict(zip(sorted_keys, random_numbers))

    return result


# function for development assistance
def output_random_experts():
    pprint.pprint("Случайно выбранные эксперты")
    pp = pprint.PrettyPrinter(sort_dicts=False)
    pp.pprint(select_random_experts())


# function for development assistance
def fill_the_cells(filepath, step):
    wb = openpyxl.load_workbook(filepath)

    sheet_name = f'Исходные данные {step} шага'
    sheet = wb[sheet_name]

    # Generate random data and fill the table
    for row_num in range(2, number_of_experts + 2):
        max_value = random.randint(8, 10)
        sheet.cell(row=row_num, column=4, value=max_value)

        avg_value = random.randint(5, 7)
        sheet.cell(row=row_num, column=3, value=avg_value)

        min_value = random.randint(1, 4)
        sheet.cell(row=row_num, column=2, value=min_value)

    # Save the workbook to a file
    wb.save(filepath)

    return pd.read_excel(filepath)


def create_source_file(filepath, step):
    wb = openpyxl.Workbook()

    default_sheet = wb.active
    wb.remove(default_sheet)

    sheet_name = f'Исходные данные {step} шага'

    sheet = wb.create_sheet(title=sheet_name)

    column_names = ['минимально', 'среднее', 'максимально']

    # generating rows with experts
    for row_num in range(1, number_of_experts + 1):
        cell_value = f'E{row_num}'
        sheet.cell(row=row_num + 1, column=1, value=cell_value)

    # generating columns with values
    for col_num, column_name in enumerate(column_names, start=1):
        sheet.cell(row=1, column=col_num + 1, value=column_name)

    # generating feedback columns
    for reason in range(len(column_names)):
        sheet.cell(row=1, column=reason + len(column_names) + 3, value=f"Объяснение {column_names[reason]}")

    wb.save(filepath)

    return pd.read_excel(filepath)


def create_calculation_sheet(filepath, step):
    wb = openpyxl.load_workbook(filepath)

    sheet_name = f'Вычисления {step} шага'
    sheet = wb.create_sheet(title=sheet_name)

    columns_names = ['Число итераций', 'Среднее оценок экспертов',
                     'Дисперсия', 'Среднеквадр. отклонение', 'Коэф. вариации', 'Асимметрия', ]

    # generating mean row
    sheet.cell(row=2, column=1, value="Среднее каждого столбца")

    # generating rows with experts
    for row_num in range(1, number_of_experts + 1):
        cell_value = f'E{row_num}'
        sheet.cell(row=row_num + 2, column=1, value=cell_value)

    # generating columns with values
    for col_num, column_name in enumerate(columns_names, start=1):
        sheet.cell(row=1, column=col_num + 1, value=column_name)

    wb.save(filepath)

    return wb, sheet


def calculations(filepath, step):
    wb, calculation_sheet = create_calculation_sheet(filepath, step)

    source_sheet_name = f'Исходные данные {step} шага'
    source_sheet = wb[source_sheet_name]

    iteration_number = 1000

    calculation_columns_names = ['Число итераций', 'Среднее оценок экспертов',
                                 'Дисперсия', 'Среднеквадр. отклонение', 'Коэф. вариации', 'Асимметрия', ]

    # заполнение поля число итераций
    for i in range(number_of_experts):
        calculation_sheet.cell(row=i + 3, column=2, value=iteration_number)

    def find_column_index_by_name(column_name):

        # find the column index based on the column name
        column_index = None
        for cell in calculation_sheet[1]:
            if cell.value == column_name:
                column_index = cell.column_letter
                return column_index

    def convert_column_letter_to_number(column_letter):
        column_number = 0
        power = 1
        for char in reversed(column_letter):
            char_value = ord(char.upper()) - ord('A') + 1
            column_number += char_value * power
            power *= 26
        return column_number

    # вычисление среднеарифметического
    def arithmetic_mean(column_name=None, row=None):

        if column_name:
            column_values = []
            column_index = find_column_index_by_name(column_name)

            if column_index:
                for cell in calculation_sheet[column_index]:
                    if isinstance(cell.value, (int, float)):
                        column_values.append(cell.value)

            if column_values:
                result = mean(column_values)
                return result, column_index

        if row:
            row_values = []
            for column in range(2, 5):
                cell = source_sheet.cell(row=row, column=column).value
                if isinstance(cell, (int, float)):
                    row_values.append(cell)

            result = mean(row_values)
            return result

        return []

    # вычисление дисперсии
    def variance(expert):
        expert_row_source_sheet = expert + 1
        min_rate = source_sheet.cell(column=2, row=expert_row_source_sheet).value
        avg_rate = source_sheet.cell(column=3, row=expert_row_source_sheet).value
        max_rate = source_sheet.cell(column=4, row=expert_row_source_sheet).value

        expert_row_calculation_sheet = expert + 2
        rates_mean = calculation_sheet.cell(column=3, row=expert_row_calculation_sheet).value

        var = ((min_rate - rates_mean) ** 2 + (avg_rate - rates_mean) ** 2 + (max_rate - avg_rate) ** 2) / 3

        return var

    # вычисление среднеквадратического отклонения
    def deviation(expert):
        expert_row_source_sheet = expert + 1
        expert_row_calculation_sheet = expert + 2

        dev = sqrt(calculation_sheet.cell(column=4, row=expert_row_calculation_sheet).value)

        return dev

    # вычисление асимметрии
    def asymmetry(expert):
        expert_row_source_sheet = expert + 1
        expert_row_calculation_sheet = expert + 2

        rates_mean = calculation_sheet.cell(column=3, row=expert_row_calculation_sheet).value
        max_rate = source_sheet.cell(column=4, row=expert_row_source_sheet).value

        dev = sqrt(calculation_sheet.cell(column=4, row=expert_row_calculation_sheet).value)

        asym = (rates_mean - max_rate) / dev

        return asym

    # вычисление коэффициента вариации
    def variation_coefficient(expert):
        expert_row_calculation_sheet = expert + 2
        deviation_number = calculation_sheet.cell(column=5, row=expert_row_calculation_sheet).value
        rates_mean = calculation_sheet.cell(column=3, row=expert_row_calculation_sheet).value

        coefficient = deviation_number / rates_mean

        return coefficient

    # 1: заполнение ячеек столбца среднее оценок экспертов
    for expert_row in range(2, number_of_experts + 2):
        column_letter = find_column_index_by_name('Среднее оценок экспертов')
        column = convert_column_letter_to_number(column_letter)

        value = arithmetic_mean(row=expert_row)
        calculation_sheet.cell(row=expert_row + 1, column=column, value=value)

    # 2: заполнение ячеек с дисперсией
    for i in range(1, number_of_experts + 1):
        calculation_sheet.cell(column=4, row=i + 2, value=variance(i))

    # 3: заполнение ячеек среднееквадр. отклонение
    for i in range(1, number_of_experts + 1):
        calculation_sheet.cell(column=5, row=i + 2, value=deviation(i))

    # 4: заполнение ячеек коэффициент вариации
    for i in range(1, number_of_experts + 1):
        calculation_sheet.cell(column=6, row=i + 2, value=variation_coefficient(i))

    # 5: заполнение ячеек асимметрии
    for i in range(1, number_of_experts + 1):
        calculation_sheet.cell(column=7, row=i + 2, value=asymmetry(i))

    # 6: среднеарифметическое для каждого слолбца
    for column in calculation_columns_names:
        result, column_index = arithmetic_mean(column_name=column)
        column_index = convert_column_letter_to_number(column_index)
        if result:
            calculation_sheet.cell(row=2, column=column_index, value=result)

    wb.save(filepath)


def main():
    source_dir = input("Выберите директорию для расположения исходного файла.\n")
    path_to_file = os.path.join(source_dir, filename)

    create_source_file(path_to_file, 1)

    fill_the_cells(path_to_file, 1)

    calculations(path_to_file, 1)
    print(get_sheet(path_to_file, 2))


main()
